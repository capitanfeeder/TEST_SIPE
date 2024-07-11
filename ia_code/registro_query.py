import pandas as pd
import os
import json
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime


def guardar_consulta_correcta(pregunta: str, consulta: str, execute: dict) -> None:
    """
    Guarda una consulta correcta en un archivo .xlsx.

    Args:
        pregunta (str): La pregunta original realizada.
        consulta (str): La consulta SQL generada.
        execute (dict): El resultado de la ejecución de la consulta.

    Returns:
        None
    """
    data = {
        "Pregunta": [pregunta],
        "Consulta": [consulta],
        "Execute": [json.dumps(execute)],
        "Datetime": [datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
    }
    df = pd.DataFrame(data)

    directorio = "./registro"
    if not os.path.isdir(directorio):
        os.makedirs(directorio)

    archivo = os.path.join(directorio, "consultas_correctas.xlsx")
    if not os.path.isfile(archivo):
        # Crear el archivo Excel y la hoja de trabajo
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Consultas correctas"

        # Escribir los encabezados de la hoja de trabajo
        ws.append(["Pregunta", "Consulta", "Execute", "Datetime"])

        # Guardar el archivo Excel
        wb.save(archivo)
    else:
        # Abrir el archivo Excel existente y la hoja de trabajo
        wb = openpyxl.load_workbook(archivo)
        ws = wb["Consultas correctas"]

        # Convertir los datos a un formato que pueda ser escrito en la hoja de trabajo
        rows = dataframe_to_rows(df, index=False, header=False)

        # Escribir los datos en la hoja de trabajo
        for row in rows:
            ws.append(row)

        # Guardar el archivo Excel
        wb.save(archivo)


def guardar_consulta_error(pregunta: str, consulta: str, error: str) -> None:
    """
    Guarda una consulta que falló en un archivo .xlsx.

    Args:
        pregunta (str): La pregunta original realizada.
        consulta (str): La consulta SQL generada.
        error (str): El mensaje de error producido.

    Returns:
        None
    """
    data = {
        "Pregunta": [pregunta],
        "Consulta": [consulta],
        "Error": [error],
        "Datetime": [datetime.datetime.now()]
    }
    df = pd.DataFrame(data)

    directorio = "./registro"
    if not os.path.isdir(directorio):
        os.makedirs(directorio)

    archivo = os.path.join(directorio, "consultas_error.xlsx")
    if not os.path.isfile(archivo):
        # Crear el archivo Excel y la hoja de trabajo
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Consultas erróneas"

        # Escribir los encabezados de la hoja de trabajo
        ws.append(["Pregunta", "Consulta", "Error", "Datetime"])

        # Guardar el archivo Excel
        wb.save(archivo)
    else:
        # Abrir el archivo Excel existente y la hoja de trabajo
        wb = openpyxl.load_workbook(archivo)
        ws = wb["Consultas erróneas"]

        # Convertir los datos a un formato que pueda ser escrito en la hoja de trabajo
        rows = dataframe_to_rows(df, index=False, header=False)

        # Escribir los datos en la hoja de trabajo
        for row in rows:
            ws.append(row)

        # Guardar el archivo Excel
        wb.save(archivo)
